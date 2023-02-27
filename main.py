from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService

from openpyxl import load_workbook
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.

driver = EdgeService(executable_path="msedgedriver.exe" )

# options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging']) 


options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)
options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36')


driver = webdriver.Edge(options=options)

# wb = load_workbook('file.xlsx')
# sheet = wb['DETAILED']
wb = openpyxl.load_workbook('file.xlsx')
sheet = wb['DETAILED']


driver.get('https://www.bing.com/')
driver.maximize_window()
driver.implicitly_wait(20)

# i = 2
# while i <= len(detailRange['A']):
#     facname = detailRange['A' + str(i)].value
    
#     try :
#         # WebDriverWait(driver,10).until(EC.visibility_of_element_located(By.XPATH('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input')))

#         driver.fi
# for row in range(2, sheet.max_row + 1):
#     # Ambil data dari kolom FACNAME
#     facname = sheet.cell(row=row, column=1).value
    
#     # Cari data di Google
#     search_input =  WebDriverWait(driver,10).until(EC.visibility_of_element_located(By.XPATH('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input')))
#     search_input.send_keys(facname + ' some other keywords')
#     search_input.submit()

#     # Tunggu hasil pencarian muncul
#     search_results = driver.find_elements_by_xpath('//div[@class="g"]')

#     # Ambil judul dan deskripsi dari hasil pencarian pertama
#     title = search_results[0].find_element_by_tag_name('h3').text
#     description = search_results[0].find_element_by_tag_name('span').text

#     # Simpan data ke dalam kolom yang berbeda
#     sheet.cell(row=row, column=2).value = title
#     sheet.cell(row=row, column=3).value = description

#     # Bersihkan input pencarian untuk pencarian berikutnya
#     search_input = driver.find_element_by_name('q')
#     search_input.clear()

for row in range(2, sheet.max_row + 1):
    # Ambil data dari kolom FACNAME
    facname = sheet.cell(row=row, column=1 ).value
    city = sheet.cell(row=row, column=3 ).value
    
    # Cari data di Google
    search_input = driver.find_element(By.NAME, 'q')
    time.sleep(5)
    search_input.send_keys(facname + ' ' , city)
    time.sleep(3)
    search_input.send_keys(Keys.ENTER)
    time.sleep(5)
    

    # search_input.send_keys(facname + ' some other keywords')
    # search_input.submit()

    # Tunggu hasil pencarian muncul
    # search_results = driver.find_elements_by_xpath('//div[@class="g"]')
    # search_results = driver.find_element(By.XPATH, '//*[@id="kp-wp-tab-overview"]/div[1]/div/div/div/div/div/div[4]/div/div/div')
    # # Ambil judul dan deskripsi dari hasil pencarian pertama
    # # Simpan data ke dalam kolom yang berbeda
    # title = search_results[0].find_element_by_tag_name('h3').text
    # sheet.cell(row=row, column=2).value = title

    # Bersihkan input pencarian untuk pencarian berikutnya
    search_input = driver.find_element(By.NAME, 'q')    
    time.sleep(2)
    search_input.clear()

wb.save('file.xlsx')
print(wb)