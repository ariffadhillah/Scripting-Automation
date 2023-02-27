from selenium import webdriver

# from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService

import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

# driver = ChromeService(executable_path="chromedriver.exe" )
driver = EdgeService(executable_path="msedgedriver.exe" )

# options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging']) 


options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)


options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.46' )

driver = webdriver.Edge(options=options)

WAIT_TIME = 60

wb = openpyxl.load_workbook('coba.xlsx')
sheet = wb['DETAILED']


driver.get('https://www.cdph.ca.gov/Programs/CHCQ/LCP/CalHealthFind/Pages/Complaint.aspx')
driver.maximize_window()
# driver.implicitly_wait(20)
time.sleep(2)
for row in range(866, 900):
# for row in range(3409, sheet.max_row + 1):
    # Ambil data dari kolom FACNAME
    facname = sheet.cell(row=row, column=1 ).value
    # city = sheet.cell(row=row, column=3 ).value
    
    # Cari data di Google
    # driver.implicitly_wait(1)
    search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
    # search_input = driver.find_element(By.ID, 'txtSearch')
    # time.sleep(1)
    search_input.send_keys(facname)
    # driver.implicitly_wait(5)
    time.sleep(2)
    search_input.send_keys(Keys.ENTER)
    # time.sleep(10)
    
    driver.implicitly_wait(1)
    try:
        facilityProviderName = driver.find_element(By.ID, 'lblFacilityName').text
        # WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblFacilityName'))).text
        streetAddress = driver.find_element(By.ID, 'lblFacilityAddress').text 
        # WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblFacilityAddress'))).text

        cityStateZIPCode = driver.find_element(By.ID, 'lblFacilityCity').text
        # WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblFacilityCity'))).text
        # facilityProviderName = WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblFacilityName'))).text
        # streetAddress = WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblFacilityAddress'))).text
        # cityStateZIPCode = WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblFacilityCity'))).text
        # label_element =  driver.find_element_by_id('lblDistrictInfo')
        # label_element = WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'lblDistrictInfo')))
        # datalblDistrictInfo = label_element.text.strip()
        # data = datalblDistrictInfo.split('\n')

        # address = data[1]
        # city_state_zip = data[2]
        # phone_number = data[3]

        print(facilityProviderName, streetAddress, cityStateZIPCode)
        time.sleep(1)
        sheet.cell(row=row, column=8).value = facilityProviderName
        sheet.cell(row=row, column=9).value = streetAddress
        sheet.cell(row=row, column=10).value = cityStateZIPCode      
        # sheet.cell(row=row, column=12).value = address
        # sheet.cell(row=row, column=13).value = city_state_zip
        # sheet.cell(row=row, column=14).value = phone_number      



    except:
        print(' ')
    
    

    # time.sleep(1)
    search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
    # time.sleep(1)
    search_input.clear()
    


    wb.save('coba.xlsx')
    print(wb)
