from selenium import webdriver
from bs4 import BeautifulSoup
# from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService

import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

# driver = ChromeService(executable_path="chromedriver.exe" )
driver = EdgeService(executable_path="msedgedriver.exe" )

options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)


options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.46' )

driver = webdriver.Edge(options=options)

WAIT_TIME = 60

driver.get('https://www.airbnb.com/rooms/39663644?category_tag=Tag%3A8099&check_in=2023-01-31&check_out=2023-02-01&source_impression_id=p3_1677139635_i4vGet1KX8CXouAd')
driver.maximize_window()


# klik tombol untuk membuka pop up
button = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="site-content"]/div/div[1]/div[1]/div[1]/div/div/div/div/section/div[2]/div[1]/span[1]/span[3]/button')))
button.click()
# button = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.XPATH, "//button[text()='142 reviews']")))
# button.click()
# time.sleep(3)


# SCROLL_PAUSE_TIME = 1

# last_height = driver.execute_script("return document.body.scrollHeight")

# while True:
#     # Scroll down to bottom
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

#     # Wait to load page
#     time.sleep(SCROLL_PAUSE_TIME)

#     # Calculate new scroll height and compare with last scroll height
#     new_height = driver.execute_script("return document.body.scrollHeight")
#     if new_height == last_height:
#         break
#     last_height = new_height

# # Scrape data from pop up
# pop_up_data = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.XPATH, ' //*[@role="dialog"]')))
# data_text = pop_up_data.text

# # Close pop up
# close_button = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Close']")))
# close_button.click()

# # Close browser
# driver.quit()

# # Process data
# print(data_text)


# close_button = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Close']")))
# close_button.click()
# print(facilityProviderName)





# scroll pop up sampai seluruh datanya terlihat
# last_height = driver.execute_script("return document.body.scrollHeight")
# while True:
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#     new_height = driver.execute_script("return document.body.scrollHeight")
#     if new_height == last_height:
#         break
#     last_height = new_height

# ambil data dari pop up
# popup = driver.find_element(By.XPATH, "//div[@class='popup']")
# soup = BeautifulSoup(popup.get_attribute('innerHTML'), 'html.parser')
# data = soup.find_all('div', {'class': 'data'})

# # print data
# for d in data:
#     print(d.text)


# tunggu hingga tombol button muncul
# button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button.l1j9v1wn.bbkw4bl.c1rxa9od.dir.dir-ltr')))
# button = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, '//*[@id="site-content"]/div/div[1]/div[4]/div/div/div/div[2]/section/div[4]/button')))
# # klik tombol button
# button.click()



# # Menunggu tombol "Pesan Sekarang" muncul dan klik tombol tersebut
# popup_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//button[contains(@data-testid, 'book-it-button')]")))
# popup_button.click()

# # Menunggu popup muncul dan ambil sumber halaman popup
# popup = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/section/div/div/div[2]/div")))
# soup = BeautifulSoup(popup.get_attribute('innerHTML'), 'html.parser')

# # soup = BeautifulSoup(driver.page_source, 'html.parser')

# # popup_html = driver.execute_script('return document.querySelector("div[data-testid=\'modalContent\']").innerHTML')

# # Parsing HTML menggunakan BeautifulSoup

# # popup = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > div:nth-child(54) > section > div > div > div._z4lmgp > div")))
# # # Ambil data HTML dari pop up
# # popup_html = popup.get_attribute('innerHTML')

# # # Parsing HTML menggunakan BeautifulSoup
# # soup = BeautifulSoup(popup_html, 'html.parser')
# # print(soup)
# # Cari semua ulasan di dalam pop up
# # reviews = soup.find_all('div', {'class': '_1s35tye2'})

# # # Looping untuk mengambil informasi dari setiap ulasan
# # for review in reviews:
# #     # Mengambil nama pengguna
# #     user_name = review.find('div', {'class': '_16shi2n'}).text.strip()
# #     # Mengambil tanggal ulasan
# #     date = review.find('div', {'class': '_1ixtnfc2'}).text.strip()
# #     # Mengambil isi ulasan
# #     review_text = review.find('span', {'class': '_1b05uo2b'}).text.strip()

# #     # Output informasi
# #     print(f"User: {user_name}")
# #     print(f"Date: {date}")
# #     print(f"Review: {review_text}")
# #     print("--------------------------------------------------")

# # # Tutup browser
# # driver.quit()



# # soup = BeautifulSoup(driver.page_source, 'html.parser')
# # # Menunggu popup muncul dan ambil sumber halaman popup
# # popup = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, "//div[@role='dialog']")))
# # # soup = BeautifulSoup(driver.page_source, 'html.parser')
# # print(soup)

# # item = soup.findAll('div', 'r1are2x1 dir dir-ltr')

# # for it in item:
# #     name = it.find('span', 'll4r2nl dir dir-ltr').text
# #     print(name)


# # title = soup.find('h3')
# # subtitle = soup.find('h2', {'data-testid': 'modal-subtitle'}).text
# # description = soup.find('div', {'data-testid': 'room-description'}).text
# # amenities = [amenity.text for amenity in soup.find_all('div', {'class': '_jfp583'})]

# # Print hasil ekstraksi
# # print("Title: ", title)
# # print("Subtitle: ", subtitle)
# # print("Description: ", description)
# # print("Amenities: ", amenities)

# # Menutup browser
# # driver.quit()

# # Mengekstrak data dari popup
# # title = soup.find('h1', {'class': '_14i3z6h'}).text
# # # subtitle = soup.find('h2', {'data-testid': 'modal-subtitle'}).text
# # # description = soup.find('div', {'data-testid': 'room-description'}).text
# # # amenities = [amenity.text for amenity in soup.find_all('div', {'class': '_jfp583'})]

# # # Print hasil ekstraksi
# # print("Title: ", title)
# # print("Subtitle: ", subtitle)
# # print("Description: ", description)
# # print("Amenities: ", amenities)

# # Menutup browser



# # tunggu hingga pop up muncul
# popup = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.CLASS_NAME, '_14vzertx')))

# # # soup = BeautifulSoup(driver.page_source, 'html.parser')
# # # reviews = soup.find_all('div', {'class': 'r1are2x1 dir dir-ltr'})
# # # for review in reviews:
# # #   print(len(review.text))


facname = 'coba'

# # # scroll ke bawah hingga semua data terlihat
popup = WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.CLASS_NAME, '_17itzz4')))

while True:
    # hitung ketinggian pop up
    height = driver.execute_script("return arguments[0].scrollHeight", popup)
    time.sleep(2)
    # scroll ke bawah
    driver.execute_script(f"arguments[0].scrollTo(0, {height});", popup)
    # driver.execute_script("window.scrollTo(0,4000)", popup)
    
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    reviews = soup.find_all('div', {'class': 'r1are2x1 dir dir-ltr'})
    # print(f"Reviews popup {+1}:")
    for review in reviews:
        print(review.text)

    # tunggu hingga data terlihat
    try:
        print('Data None')
        # WebDriverWait(driver, WAIT_TIME).until(EC.presence_of_element_located((By.CSS_SELECTOR, '_17itzz4')))
    except:
        break

# ambil semua data dari pop up
data = popup.text

# tampilkan data
print(data)

# tutup browser
driver.quit()


# # button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "l1j9v1wn")))

# # # klik tombol
# # button.click()

# # wait = WebDriverWait(driver, 2)
# # show_all_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-section-id='REVIEWS_DEFAULT']")))
# # show_all_button.click()

# # # Wait for the popup to appear and extract the data
# # popup = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@data-section-id='REVIEWS_DEFAULT']")))
# # print(popup.text)

# # # Close the popup and the driver
# # close_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Close']")))
# # close_button.click()
# # driver.quit()

# # https://www.airbnb.com/rooms/39663644?category_tag=Tag%3A8099&check_in=2023-01-31&check_out=2023-02-01&source_impression_id=p3_1677139635_i4vGet1KX8CXouAd


# # for row in range(4286, 4289):
# #     facname = sheet.cell(row=row, column=1 ).value
# #     search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
# #     search_input.send_keys(facname)
# #     time.sleep(2)
# #     search_input.send_keys(Keys.ENTER)
# #     try:
# #         facilityProviderName = driver.find_element(By.ID, 'lblFacilityName').text
# #         streetAddress = driver.find_element(By.ID, 'lblFacilityAddress').text 
# #         cityStateZIPCode = driver.find_element(By.ID, 'lblFacilityCity').text
# #         print(facilityProviderName, streetAddress, cityStateZIPCode)
# #         time.sleep(1)
# #         sheet.cell(row=row, column=8).value = facilityProviderName
# #         sheet.cell(row=row, column=9).value = streetAddress
# #         sheet.cell(row=row, column=10).value = cityStateZIPCode      
# #     except:
# #         print(' ')

# #     search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
# #     search_input.clear()
# #     wb.save('Combined - HHAs & Hospices - addresses only-1.xlsx')





