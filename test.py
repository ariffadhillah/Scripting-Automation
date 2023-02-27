from selenium import webdriver

from selenium.webdriver.edge.service import Service as EdgeService

import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


driver = EdgeService(executable_path="msedgedriver.exe" )

# options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging']) 


options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)
options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36')


driver = webdriver.Edge(options=options)


wb = openpyxl.load_workbook('Combined - HHAs & Hospices - addresses only - Copy (2).xlsx')
sheet = wb['DETAILED']


driver.get('https://www.bing.com/')
driver.maximize_window()
driver.implicitly_wait(20)

for row in range(2, sheet.max_row + 1):
    # Ambil data dari kolom FACNAME
    facname = sheet.cell(row=row, column=1 ).value
    city = sheet.cell(row=row, column=3 ).value
    
    # Cari data di Google
    search_input = driver.find_element(By.NAME, 'q')
    time.sleep(1)
    search_input.send_keys(facname + ' ' , city + ' ' , ' npi usa')
    time.sleep(1)
    search_input.send_keys(Keys.ENTER)
    # time.sleep(10)
    
    driver.implicitly_wait(2)
    try:
        intro_flex = driver.find_element(By.CLASS_NAME, 'ntro-vf-flex')
        npi = intro_flex.find_element(By.XPATH, './/span[text()="NPI: "]/following-sibling::span').text
        provider_name = intro_flex.find_element(By.XPATH, './/span[text()="Provider Name: "]/following-sibling::span').text
        provider_location_address = intro_flex.find_element(By.XPATH, './/span[text()="Provider Location Address: "]/following-sibling::span').text
        # provider_mailing_address = intro_flex.find_element(By.XPATH, './/span[text()="Provider Mailing Address: "]/following-sibling::span').text
        
        print(npi, provider_name, provider_location_address)
        sheet.cell(row=row, column=8).value = npi
        sheet.cell(row=row, column=9).value = provider_name
        sheet.cell(row=row, column=10).value = provider_location_address

    except:
        print(' ')
    
    

    # time.sleep(1)
    search_input = driver.find_element(By.NAME, 'q')
    time.sleep(1)
    search_input.clear()
    


    wb.save('Combined - HHAs & Hospices - addresses only - Copy (2).xlsx')
    print(wb)
