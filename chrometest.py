from selenium import webdriver

# from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService

import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


# driver = ChromeService(executable_path="chromedriver.exe" )
driver = EdgeService(executable_path="chromedriver.exe" )

# options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging']) 


options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)


options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.46' )


driver = webdriver.Edge(options=options)



WAIT_TIME = 300

wb = openpyxl.load_workbook('Combined - HHAs & Hospices - addresses only-.xlsx')
sheet = wb['DETAILED']


driver.get('https://www.cdph.ca.gov/Programs/CHCQ/LCP/CalHealthFind/Pages/SearchResult.aspx')
driver.maximize_window()
driver.implicitly_wait(20)

for row in range(2, sheet.max_row + 1):
    # Ambil data dari kolom FACNAME
    facname = sheet.cell(row=row, column=7 ).value
    city = sheet.cell(row=row, column=3 ).value
    
    # Cari data di Google
    search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
    # search_input = driver.find_element(By.ID, 'txtSearch')
    # time.sleep(1)
    search_input.send_keys(facname)
    time.sleep(1)
    search_input.send_keys(Keys.ENTER)
    # time.sleep(10)
    
    driver.implicitly_wait(2)
    # intro_flex = driver.find_element(By.CLASS_NAME, 'ntro-vf-flex')

    # try:
    clicktitlethumbInfo = WebDriverWait(driver,WAIT_TIME).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="scrollContainerResults"]/div/div/div[3]/div[2]/h3'))).click()
        # thumbContainerHover0
    thumbInfo = driver.find_element(By.ID, 'thumbContainerHover0')

    view_details = WebDriverWait(thumbInfo, WAIT_TIME).until(EC.element_to_be_clickable((By.CLASS_NAME, '//a[text()="File a Complaint"]'))).click()
    # closeAutoPopup = WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.XPATH, '//a[contains(text(), "Matikan untuk: Inggris")]')))
    # closeAutoPopup.click()
    # closeAutoPopup = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.XPATH, '//*[@id=":2.close"]'))).click()
    driver.implicitly_wait(60)
    time.sleep(10)



       
        # myModal = driver.find_element(By.ID, 'myModal')
        # try:
        # except:
        #     None

        # name = myModal.find_element(By.XPATH, 'lblFacType').text
        # print(name)
        # lblFacType
        # view_details = WebDriverWait(thumbInfo, WAIT_TIME).until(EC.element_to_be_clickable((By.XPATH, '//a[text()="View Details"]'))).click()



    clicktitlethumbInfo = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.CLASS_NAME, 'popupClose')))
    driver.execute_script("arguments[0].click();", clicktitlethumbInfo)


    # try:
        # intro_flex = driver.find_element(By.CLASS_NAME, 'thumbInfo')
        # title = intro_flex.find_element(By.XPATH, '//*[@id="scrollContainerResults"]/div/div/div[3]/div[2]/h3/span').click()
        



        # modalpopup = driver.find_element(By.ID, 'myModal')
        
        # btnClose = modalpopup.find_element(By.CLASS_NAME, 'popupClose').click()
        # thumbInfo.click()
        # driver.find_element(By.XPATH, '//*[@id="scrollContainerResults"]/div/div/div[3]/div[2]/h3/span').click()
        # print(title)
        
        # npi = intro_flex.find_element(By.XPATH, './/span[text()="NPI: "]/following-sibling::span').text
        # provider_name = intro_flex.find_element(By.XPATH, './/span[text()="Provider Name: "]/following-sibling::span').text
        # provider_location_address = intro_flex.find_element(By.XPATH, './/span[text()="Provider Location Address: "]/following-sibling::span').text
        # # provider_mailing_address = intro_flex.find_element(By.XPATH, './/span[text()="Provider Mailing Address: "]/following-sibling::span').text
        
        # print(npi, provider_name, provider_location_address)
        # sheet.cell(row=row, column=8).value = npi
        # sheet.cell(row=row, column=9).value = provider_name
        # sheet.cell(row=row, column=10).value = provider_location_address

    # except:
    #     print(' data none ')
    
    

    time.sleep(1)
    search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
    time.sleep(1)
    search_input.clear()
    


    wb.save('Combined - HHAs & Hospices - addresses only-.xlsx')
    print(wb)
