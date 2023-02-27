from selenium import webdriver

# from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService

import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

# driver = ChromeService(executable_path="chromedriver.exe" )
driver = EdgeService(executable_path="msedgedriver.exe" )

# options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging']) 


options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)


options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.46' )

driver = webdriver.Edge(options=options)

WAIT_TIME = 60

wb = openpyxl.load_workbook('Combined - HHAs & Hospices - addresses only-1.xlsx')
sheet = wb['DETAILED']


driver.get('https://www.cdph.ca.gov/Programs/CHCQ/LCP/CalHealthFind/Pages/Complaint.aspx')
driver.maximize_window()
# driver.implicitly_wait(20)
# time.sleep(2)
# 3961 - 3973
# 4827 - 4830
# 4835 - 4837 
# 4866 
# 4871  
# 4873  
# 4956  
# 4958  
# 4973  
# 4993  
# 4995  
# 5001  
# 5032  
# 5036  
# 5046  
# 5048  
# 5051 - 5053    
# 5063 - 5065    
# 5063 - 5065    
# 5083 - 5085    
# 5091    
# 5102 - 5105 
# 20 - 22 
# 44 
# 47 
# 18-396
for row in range(4286, 4289):
    facname = sheet.cell(row=row, column=1 ).value
    search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
    search_input.send_keys(facname)
    time.sleep(2)
    search_input.send_keys(Keys.ENTER)
    try:
        facilityProviderName = driver.find_element(By.ID, 'lblFacilityName').text
        streetAddress = driver.find_element(By.ID, 'lblFacilityAddress').text 
        cityStateZIPCode = driver.find_element(By.ID, 'lblFacilityCity').text
        print(facilityProviderName, streetAddress, cityStateZIPCode)
        time.sleep(1)
        sheet.cell(row=row, column=8).value = facilityProviderName
        sheet.cell(row=row, column=9).value = streetAddress
        sheet.cell(row=row, column=10).value = cityStateZIPCode      
    except:
        print(' ')

    search_input = WebDriverWait(driver, WAIT_TIME).until(EC.element_to_be_clickable((By.ID, 'txtSearch')))
    search_input.clear()
    wb.save('Combined - HHAs & Hospices - addresses only-1.xlsx')

driver.quit()
